import sys
import openpyxl
import json  # Добавляем импорт модуля json


# Проверка аргументов командной строки
if len(sys.argv) < 2:
    print("Usage: python parse.py <filename>")
    sys.exit(1)

# Получение имени файла из аргументов командной строки
filename = sys.argv[1]
file_path = "asset/" + filename + ".xlsx"

print("Чтение из файла: " + file_path)

# Открываем файл Excel
workbook = openpyxl.load_workbook(file_path)

# Получаем активный лист
sheet = workbook.active

data = []

# Проходим по строкам
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    # Получаем нужные данные по индексам
    client_tariff = row[0].value 
    client_type = row[1].value  
    client_name = row[2].value  
    client_company_type = row[3].value
    client_company_department = row[4].value
    manager_name = row[5].value
    client_sex = row[6].value
    client_area = row[7].value
    client_firsname = row[8].value
    client_lastname = row[10].value
    client_timezone = row[11].value
    
    email = row[26].value     
    phone = row[27].value      

        # Создание словаря для строки
    client_data = {
        'client_tariff':client_tariff,
        'client_name': client_name,
        'client_type': client_type,
        'client_company_type': client_company_type,
        'client_company_department': client_company_department,
        'manager_name': manager_name,
        'client_sex': client_sex,
        'client_firstname': client_firsname,
        'client_lastname': client_lastname,
        'client_timezone': client_timezone,
        'email': email,
        'phone': "+" + str(phone)
    }

    data.append(client_data)

    # Преобразование списка данных в формат JSON
    json_data = json.dumps(data, ensure_ascii=False, indent=4)

    # Запись в файл JSON (по желанию)
    outputfile = "json/" + filename + ".json"
    print("Запись в файл: " + outputfile)
    with open(outputfile, 'w', encoding='utf-8') as outputfile: outputfile.write(json_data)
